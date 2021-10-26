from contracts import contract
from openpyxl import load_workbook
from datetime import datetime
import threading
import logging
import numpy as np
import time
import ui
import sys
import os

near_array = [[0] * 12 for _ in range(15)]
far_array = [[0] * 12 for _ in range(15)]

now = datetime.now()

coil_equal = {
    '0': 0,
    '1': 3,
    '2': 6,
    '3': 9
}

row_equal = {
    '0': 40,
    '1': 55
}

array_equal = {
    '0': near_array,
    '1': far_array
}

interface = ui.UInterface()


def except_log():
    exc_type, exc_obj, exc_tb = sys.exc_info()
    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]

    f1 = open("log.ini", "a")
    f1.write("[" + now.strftime("%d/%m/%Y %H:%M:%S") + "]" + exc_type + fname + exc_tb.tb_lineno + "\n")
    f1.close()


try:
    interface.welcome_output()
    txt_file = interface.txt_path_input()
    interface.xls_path_input()
    xls_file = interface.tor_number_input()
except Exception as ui_argument:
    except_log()


def extract_from_file(file_path):
    # Open file and read it data
    file = open(file_path)
    file_data = file.read()

    # Split string to list and initialize it as numpy array
    file_data = file_data.split()
    file_data = np.array(file_data)
    # Convert array data to int type
    file_data = file_data.astype(int)

    # Serialize necessary data from array
    row_pos, coil_pos, coil_number = file_data[[4, 5, 6]]
    cos_sin_phase = np.delete(file_data, [2, 4, 5, 6])

    return cos_sin_phase, row_pos, coil_pos, coil_number


def put_to_array(cos_sin_phase, row_pos, coil_number, array):
    column_pos = coil_equal[str(coil_number)]

    for value_amount, value in enumerate(cos_sin_phase):
        array[row_pos][value_amount + column_pos] = value


def put_to_excel():
    try:
        global xls_file

        book = load_workbook(xls_file)
        sheet = book.active
        prev_data, prev_coil_pos = list(), list()

        while True:
            time.sleep(0.15)
            cos_sin_phase, row_pos, coil_pos, coil_number = extract_from_file(str(txt_file))

            if not np.array_equal(prev_data, cos_sin_phase):

                put_to_array(cos_sin_phase, row_pos, coil_number, array_equal[str(coil_pos)])
                interface.generate_table(array_equal[str(coil_pos)], coil_pos)

                for index, value in enumerate(array_equal[str(coil_pos)][row_pos]):
                    sheet.cell(row=row_equal[str(coil_pos)] + row_pos, column=(17 + index)).value = value
                    book.save(xls_file)
                prev_data = cos_sin_phase

    except Exception as excel_argument:
        except_log()


t1 = threading.Thread(target=put_to_excel)
t1.start()
